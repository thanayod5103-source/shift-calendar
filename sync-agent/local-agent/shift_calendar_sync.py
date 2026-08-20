import base64, hashlib, json, os, sys, time
from datetime import datetime, timezone
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CONFIG = ROOT / 'config.json'
STATE = ROOT / 'state.json'

def load_json(path, default=None):
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return {} if default is None else default

def iso_now():
    return datetime.now(timezone.utc).isoformat()

def cell_rgb(cell):
    fg = cell.fill.fgColor
    rgb = (fg.rgb or '').replace('FF','',1).upper() if fg.type == 'rgb' else ''
    return {'rgb': rgb, 'type': fg.type, 'theme': fg.theme, 'tint': fg.tint}

def normalize(cell):
    raw = '' if cell.value is None else str(cell.value).strip()
    up = raw.upper()
    if up in ('D','DAY'): status = 'Day'
    elif up in ('N','NIGHT'): status = 'Night'
    elif up in ('OFF','O','พัก','OFF DUTY'): status = 'Off'
    elif any(k in up for k in ('TRAIN','อบรม','IDP','FIRE FIGHT')): status = 'Training'
    else: status = raw
    fill = cell_rgb(cell)
    rgb = fill['rgb']
    ot = (status == 'Day' and rgb in ('385724','548235','274E13')) or (status == 'Night' and rgb in ('002060','083B82','1F4E78'))
    return {'raw': raw, 'status': status, 'ot': ot, 'fill': fill}

def parse_book(path):
    wb = load_workbook(path, data_only=False, read_only=False)
    months = {}
    wanted = {f'2026_{m}' for m in ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']}
    for sheet_name in wb.sheetnames:
        if sheet_name not in wanted:
            continue
        ws = wb[sheet_name]
        start_col = next((c for c in range(1, ws.max_column + 1) if ws.cell(3,c).value == 1), None)
        if not start_col:
            raise RuntimeError(f'{sheet_name}: day 1 header not found')
        people, current_shift = [], None
        for r in range(1, ws.max_row + 1):
            marker = str(ws.cell(r,1).value or '').strip()
            namev = str(ws.cell(r,2).value or '').strip()
            position = str(ws.cell(r,3).value or '').strip()
            empid = str(ws.cell(r,5).value or '').strip()
            marker_up = marker.upper().replace('\n',' ')
            if marker_up in ('A','B','C','D','T'):
                current_shift = marker_up
            elif 'TRUCK' in marker_up:
                current_shift = 'T'
            if not current_shift or not namev or not empid:
                continue
            text = (namev + ' ' + position).lower()
            if any(x in text for x in ('signature','approval','approved','prepared','ลงชื่อ')):
                continue
            schedule = {}
            for day in range(1,32):
                cell = ws.cell(r, start_col + day - 1)
                if cell.value not in (None,''):
                    schedule[str(day)] = normalize(cell)
            people.append({'shift': current_shift, 'name': namev, 'position': position, 'id': empid, 'schedule': schedule})
        months[sheet_name] = people
    if len(months) != 12:
        raise RuntimeError(f'Expected 12 production sheets, found {len(months)}')
    if not any(months.values()):
        raise RuntimeError('Workbook parsed but no personnel were found')
    return {'generatedAt': iso_now(), 'source': 'local-sync-agent', 'months': months}

def github_put(repo, branch, path, text, token, message):
    url = f'https://api.github.com/repos/{repo}/contents/{path}'
    h = {'Authorization': f'Bearer {token}', 'Accept': 'application/vnd.github+json'}
    old = requests.get(url, headers=h, params={'ref': branch}, timeout=30)
    if old.status_code not in (200,404):
        raise RuntimeError(f'GitHub read failed for {path}: {old.status_code} {old.text[:300]}')
    sha = old.json().get('sha') if old.status_code == 200 else None
    payload = {'message': message, 'content': base64.b64encode(text.encode()).decode(), 'branch': branch}
    if sha: payload['sha'] = sha
    r = requests.put(url, headers=h, json=payload, timeout=30)
    if not r.ok:
        raise RuntimeError(f'GitHub write failed for {path}: {r.status_code} {r.text[:500]}')

def run_once(cfg, force=False):
    path = Path(cfg['workbook_path'])
    if not path.exists():
        raise FileNotFoundError(f'Workbook not found: {path}')
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    state = load_json(STATE,{})
    if not force and state.get('sha256') == digest:
        print('No workbook change')
        return False
    data = parse_book(path)
    meta = {'source': path.name, 'sha256': digest, 'lastModifiedLocal': datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat(), 'syncedAt': iso_now(), 'agent': 'local-windows', 'status': 'ok', 'months': len(data['months'])}
    token = os.environ.get('GITHUB_TOKEN')
    if not token:
        raise RuntimeError('GITHUB_TOKEN is missing. Set it before starting the agent.')
    github_put(cfg['repository'], cfg.get('branch','main'), cfg.get('output_path','data/schedule.json'), json.dumps(data,ensure_ascii=False,separators=(',',':')), token, 'sync: update shift schedule from local Excel agent')
    github_put(cfg['repository'], cfg.get('branch','main'), cfg.get('metadata_path','data/metadata.json'), json.dumps(meta,ensure_ascii=False,indent=2), token, 'sync: update shift schedule metadata')
    STATE.write_text(json.dumps({'sha256':digest,'syncedAt':iso_now()},indent=2),encoding='utf-8')
    print(f'Sync completed: {len(data["months"])} months uploaded')
    return True

def main():
    cfg = load_json(CONFIG)
    if not cfg:
        raise RuntimeError('Copy config.example.json to config.json and set workbook_path')
    if '--once' in sys.argv:
        run_once(cfg, force=True)
        return
    interval = max(60,int(cfg.get('interval_seconds',300)))
    print(f'Local Shift Calendar Sync Agent started. Checking every {interval} seconds.')
    while True:
        try: run_once(cfg)
        except Exception as e: print(f'Sync error: {e}',file=sys.stderr)
        time.sleep(interval)
if __name__ == '__main__':
    main()
